import pandas as pd
import requests
from bs4 import BeautifulSoup
import pandas as pd
import nltk
from textblob import TextBlob
from textstat import flesch_reading_ease, flesch_kincaid_grade, syllable_count, lexicon_count
nltk.download('averaged_perceptron_tagger')
import openpyxl

df = pd.read_excel('input.xlsx')
urls = df['URL'].tolist()
ids = df['URL_ID'].tolist()

id = [int(x) for x in ids]

for i in range(len(urls)):
    response = requests.get(urls[i])
    soup = BeautifulSoup(response.content, 'html.parser')
    article = soup.find('article')
    if article is not None:
        title_elem = article.find('h1')
        if title_elem is not None:
            title = title_elem.get_text().strip()
        else:
            title_elem = article.find('div', {'class': 'tdb-title-text'})
            if title_elem is not None:
                title = title_elem.get_text().strip()
            else:
                title = "No title found"
        text_elem = article.find_all('p')
        text = ""
        for p in text_elem:
            if text_elem is not None:
                text += p.get_text().strip() + "\n"
            else:
                text += "No text found"
        filename = f"{id[i]}.txt"
        with open(filename, 'w',encoding='utf-8') as f:
            f.write(title + '\n\n' + text)
    else:
        print(f"No article found for URL {urls[i]}")


workbook = openpyxl.load_workbook('Output Data Structure.xlsx')
sheet = workbook['Sheet1']
row = 2
# Loop through each file from 1.txt to n.txt
for i in range(37, 151):
    filename = f"{i}.txt"

    try:
        with open(filename, 'r', encoding='utf-8') as f:
            text = f.read()

        # Perform analysis on the text

        # Append the results to the respective lists

    except FileNotFoundError:
        print(f"{filename} not found, moving to next file.")
        sheet.cell(row=row, column=3).value = ""
        sheet.cell(row=row, column=4).value = ""
        sheet.cell(row=row, column=5).value = ""
        sheet.cell(row=row, column=6).value = ""
        sheet.cell(row=row, column=7).value = ""
        sheet.cell(row=row, column=8).value = ""
        sheet.cell(row=row, column=9).value = ""
        sheet.cell(row=row, column=10).value = ""
        sheet.cell(row=row, column=11).value = ""
        sheet.cell(row=row, column=12).value = ""
        sheet.cell(row=row, column=13).value = ""
        sheet.cell(row=row, column=14).value = ""
        sheet.cell(row=row, column=15).value = ""
        row += 1
        continue

    # Read the file and store the text
    with open(filename, 'r', encoding='utf-8') as f:
        text = f.read()

    # Perform analysis on the text using NLTK and Textstat libraries
    sentences = nltk.sent_tokenize(text)
    words = nltk.word_tokenize(text)
    personal_pronoun_count = sum(1 for word, pos in nltk.pos_tag(words) if pos == 'PRP' or pos == 'PRP$')

    positive_score = 0
    negative_score = 0
    polarity_score = 0
    subjectivity_score = 0
    avg_sentence_length = 0
    percent_complex_words = 0
    fog_index = 0
    avg_words_per_sentence = 0
    complex_word_count = 0
    word_count = len(words)
    syllables_per_word_val = syllable_count(text) / word_count
    personal_pronouns_val = personal_pronoun_count
    avg_word_lengths_val = sum(len(word) for word in words) / word_count

    for sentence in sentences:
        sentence_words = nltk.word_tokenize(sentence)
        sentence_word_count = len(sentence_words)
        sentence_complex_word_count = sum(1 for word in sentence_words if syllable_count(word) >= 3)

        positive_score += sum(1 for word in sentence_words if TextBlob(word).sentiment.polarity > 0)
        negative_score += sum(1 for word in sentence_words if TextBlob(word).sentiment.polarity < 0)
        polarity_score += TextBlob(sentence).sentiment.polarity
        subjectivity_score += TextBlob(sentence).sentiment.subjectivity
        avg_sentence_length += sentence_word_count
        percent_complex_words += sentence_complex_word_count / sentence_word_count
        complex_word_count += sentence_complex_word_count

    if len(sentences) > 0:
        avg_sentence_length /= len(sentences)
        avg_words_per_sentence = word_count / len(sentences)
    if word_count > 0:
        percent_complex_words *= 100 / len(sentences)
        fog_index = 0.4 * (avg_words_per_sentence + percent_complex_words)

    # Append the results to the respective lists
    print(positive_score)
    print(negative_score)
    print(polarity_score)
    print(subjectivity_score)
    print(avg_sentence_length)
    print(percent_complex_words)
    print(fog_index)
    print(avg_words_per_sentence)
    print(complex_word_count)
    print(word_count)
    print(syllables_per_word_val)
    print(avg_word_lengths_val)
    print(personal_pronouns_val)

    # Add the data to the DataFrame
    sheet.cell(row=row, column=3).value = positive_score
    sheet.cell(row=row, column=4).value = negative_score
    sheet.cell(row=row, column=5).value = polarity_score
    sheet.cell(row=row, column=6).value = subjectivity_score
    sheet.cell(row=row, column=7).value = avg_sentence_length
    sheet.cell(row=row, column=8).value = percent_complex_words
    sheet.cell(row=row, column=9).value = fog_index
    sheet.cell(row=row, column=10).value = avg_words_per_sentence
    sheet.cell(row=row, column=11).value = complex_word_count
    sheet.cell(row=row, column=12).value = word_count
    sheet.cell(row=row, column=13).value = syllables_per_word_val
    sheet.cell(row=row, column=14).value = personal_pronouns_val
    sheet.cell(row=row, column=15).value = avg_word_lengths_val
    row += 1

workbook.save('Output Data Structure.xlsx')